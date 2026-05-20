#Implementación y Análisis Asintótico de Algoritmos de Ordenamiento

import sqlite3
import time
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



def bubble_sort(datos):
    arr = datos[:]  # copia para no modificar el original
    n = len(arr)
    for i in range(n - 1):
        intercambiado = False
        for j in range(n - 1 - i):
            if arr[j][2] > arr[j + 1][2]:          # compara puntaje_evaluacion
                arr[j], arr[j + 1] = arr[j + 1], arr[j]
                intercambiado = True
        if not intercambiado:   # optimización: detiene si ya está ordenado
            break
    return arr


def insertion_sort(datos):
    arr = datos[:]
    for i in range(1, len(arr)):
        clave = arr[i]
        j = i - 1
        while j >= 0 and arr[j][2] > clave[2]:     # compara puntaje_evaluacion
            arr[j + 1] = arr[j]
            j -= 1
        arr[j + 1] = clave
    return arr


def selection_sort(datos):
    arr = datos[:]
    n = len(arr)
    for i in range(n):
        idx_min = i
        for j in range(i + 1, n):
            if arr[j][2] < arr[idx_min][2]:         # compara puntaje_evaluacion
                idx_min = j
        arr[i], arr[idx_min] = arr[idx_min], arr[i]
    return arr



def cargar_desde_bd(ruta_db="personas.db"):
    if not os.path.exists(ruta_db):
        print(f"  [!] No se encontró el archivo '{ruta_db}'.")
        return None
    try:
        conn = sqlite3.connect(ruta_db)
        cursor = conn.cursor()
        cursor.execute("SELECT id_persona, nombre, edad, puntaje_evaluacion FROM personas")
        datos = cursor.fetchall()
        conn.close()
        return datos
    except Exception as e:
        print(f"  [!] Error al leer la base de datos: {e}")
        return None


def cargar_desde_txt(ruta):
    try:
        datos = []
        with open(ruta, encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if not linea or linea.startswith("#"):
                    continue
                partes = linea.split(",")
                datos.append((int(partes[0]), partes[1].strip(),
                              int(partes[2]), int(partes[3])))
        return datos
    except Exception as e:
        print(f"  [!] Error al leer el archivo: {e}")
        return None


def cargar_desde_csv(ruta):
    try:
        import csv
        datos = []
        with open(ruta, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for fila in reader:
                datos.append((int(fila["id_persona"]), fila["nombre"],
                              int(fila["edad"]), int(fila["puntaje_evaluacion"])))
        return datos
    except Exception as e:
        print(f"  [!] Error al leer el CSV: {e}")
        return None


COMPLEJIDADES = {
    "Bubble Sort": {
        "peor":    "O(n²)",
        "promedio": "Θ(n²)",
        "mejor":   "Ω(n)",
        "descripcion": (
            "Compara pares adyacentes y los intercambia si están fuera de orden. "
            "En el mejor caso (datos ya ordenados) solo hace n comparaciones gracias "
            "a la bandera de intercambio, logrando Ω(n). En los demás casos, "
            "su doble bucle anidado produce O(n²)."
        ),
    },
    "Insertion Sort": {
        "peor":    "O(n²)",
        "promedio": "Θ(n²)",
        "mejor":   "Ω(n)",
        "descripcion": (
            "Toma cada elemento y lo inserta en su posición correcta dentro "
            "de la parte ya ordenada. Cuando el arreglo está ordenado, cada "
            "inserción cuesta O(1), dando Ω(n) en total. Para datos invertidos "
            "o aleatorios el costo sube a O(n²)."
        ),
    },
    "Selection Sort": {
        "peor":    "O(n²)",
        "promedio": "Θ(n²)",
        "mejor":   "Ω(n²)",
        "descripcion": (
            "Busca el mínimo en el subarreglo no ordenado y lo coloca al frente. "
            "Siempre recorre todo el subarreglo sin importar el estado inicial, "
            "por lo que su complejidad es Ω(n²) incluso en el mejor caso. "
            "No se beneficia de datos parcialmente ordenados."
        ),
    },
}


def exportar_excel(datos_originales, resultados, ruta_salida="reporte_ordenamiento.xlsx"):
    wb = openpyxl.Workbook()

    COLOR_HEADER    = "1F3864"
    COLOR_SUBHEADER = "2E75B6"
    COLOR_FILA_PAR  = "D9E1F2"
    COLOR_VERDE     = "375623"
    COLOR_AMARILLO  = "7F6000"
    COLOR_ROJO      = "843C0C"

    def header_font(color="FFFFFF", bold=True, size=11):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def data_font(bold=False, size=10):
        return Font(name="Arial", bold=bold, size=size)

    def fill(color):
        return PatternFill("solid", start_color=color, fgColor=color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def borde():
        lado = Side(style="thin", color="BFBFBF")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    def escribir_tabla_personas(ws, datos, titulo):
        ws.merge_cells("A1:D1")
        ws["A1"] = titulo
        ws["A1"].font = header_font(size=13)
        ws["A1"].fill = fill(COLOR_HEADER)
        ws["A1"].alignment = center()

        encabezados = ["ID", "Nombre", "Edad", "Puntaje Evaluación"]
        anchos = [8, 28, 10, 22]
        for col, (enc, ancho) in enumerate(zip(encabezados, anchos), 1):
            celda = ws.cell(row=2, column=col, value=enc)
            celda.font = header_font()
            celda.fill = fill(COLOR_SUBHEADER)
            celda.alignment = center()
            celda.border = borde()
            ws.column_dimensions[get_column_letter(col)].width = ancho

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 20

        for fila_idx, persona in enumerate(datos, start=3):
            color_fila = COLOR_FILA_PAR if fila_idx % 2 == 0 else "FFFFFF"
            valores = [persona[0], persona[1], persona[2], persona[3]]
            for col, val in enumerate(valores, 1):
                celda = ws.cell(row=fila_idx, column=col, value=val)
                celda.font = data_font()
                celda.fill = fill(color_fila)
                celda.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                            vertical="center")
                celda.border = borde()

        ws.freeze_panes = "A3"

    ws_orig = wb.active
    ws_orig.title = "Datos Originales"
    escribir_tabla_personas(ws_orig, datos_originales,
                            f" Datos Originales  ({len(datos_originales):,} registros)")

    for nombre_algo, info in resultados.items():
        ws = wb.create_sheet(title=nombre_algo[:31])
        tiempo_ms = info["tiempo_ms"]
        datos_ord = info["datos"]
        escribir_tabla_personas(
            ws, datos_ord,
            f" {nombre_algo}  |  Tiempo: {tiempo_ms:.2f} ms  |  "
            f"{len(datos_ord):,} registros ordenados por puntaje ↑"
        )

    wb.save(ruta_salida)
    return ruta_salida


def separador():
    print("─" * 58)

def mostrar_menu():
    separador()
    print("  TALLER: ANÁLISIS ASINTÓTICO DE ALGORITMOS")
    separador()
    print("  [1] Cargar datos (Base de Datos / TXT / CSV)")
    print("  [2] Ordenar usando Bubble Sort    (Burbuja)")
    print("  [3] Ordenar usando Insertion Sort (Inserción)")
    print("  [4] Ordenar usando Selection Sort (Selección)")
    print("  [5] Exportar reporte Excel")
    print("  [6] Salir")
    separador()

def ejecutar_ordenamiento(nombre, funcion, datos):
    print(f"\n  Ejecutando {nombre} sobre {len(datos):,} registros...")
    inicio = time.perf_counter()
    resultado = funcion(datos)
    fin = time.perf_counter()
    tiempo_ms = (fin - inicio) * 1000

    comp = COMPLEJIDADES[nombre]
    print(f"\n  Ordenamiento completado.")
    print(f"   Tiempo de ejecución : {tiempo_ms:.4f} ms  ({tiempo_ms/1000:.6f} s)")
    separador()
    print(f"  Complejidad teórica de {nombre}:")
    print(f"     • Peor caso   (Big O)  : {comp['peor']}")
    print(f"     • Caso promedio (Big Θ): {comp['promedio']}")
    print(f"     • Mejor caso  (Big Ω)  : {comp['mejor']}")
    separador()
    return resultado, tiempo_ms

def main():
    datos = None
    resultados = {}

    while True:
        mostrar_menu()
        opcion = input("  Seleccione una opción: ").strip()

        if opcion == "1":
            print("\n  ¿Fuente de datos?")
            print("  [1] Base de datos SQLite (personas.db)")
            print("  [2] Archivo .txt")
            print("  [3] Archivo .csv")
            fuente = input("  Opción: ").strip()

            if fuente == "1":
                ruta = input("  Ruta del .db [Enter = personas.db]: ").strip() or "personas.db"
                datos = cargar_desde_bd(ruta)
            elif fuente == "2":
                ruta = input("  Ruta del .txt: ").strip()
                datos = cargar_desde_txt(ruta)
            elif fuente == "3":
                ruta = input("  Ruta del .csv: ").strip()
                datos = cargar_desde_csv(ruta)
            else:
                print("  [!] Opción no válida.")
                continue

            if datos:
                print(f"\n  {len(datos):,} registros cargados exitosamente.")
            else:
                print("  [!] No se pudieron cargar los datos.")

        elif opcion == "2":
            if not datos:
                print("  [!] Primero cargue los datos (opción 1).")
            else:
                res, ms = ejecutar_ordenamiento("Bubble Sort", bubble_sort, datos)
                resultados["Bubble Sort"] = {"datos": res, "tiempo_ms": ms}

        elif opcion == "3":
            if not datos:
                print("  [!] Primero cargue los datos (opción 1).")
            else:
                res, ms = ejecutar_ordenamiento("Insertion Sort", insertion_sort, datos)
                resultados["Insertion Sort"] = {"datos": res, "tiempo_ms": ms}

        elif opcion == "4":
            if not datos:
                print("  [!] Primero cargue los datos (opción 1).")
            else:
                res, ms = ejecutar_ordenamiento("Selection Sort", selection_sort, datos)
                resultados["Selection Sort"] = {"datos": res, "tiempo_ms": ms}

        elif opcion == "5":
            if not datos:
                print("  [!] Primero cargue los datos (opción 1).")
            elif not resultados:
                print("  [!] Ejecute al menos un algoritmo antes de exportar.")
            else:
                salida = input("  Nombre del archivo Excel [Enter = reporte_ordenamiento.xlsx]: ").strip()
                salida = salida or "reporte_ordenamiento.xlsx"
                ruta_final = exportar_excel(datos, resultados, salida)
                print(f"\n  Reporte exportado: {ruta_final}")

        elif opcion == "6":
            print("\n  ¡Hasta luego!\n")
            break
        else:
            print("  [!] Opción no válida, intente de nuevo.")

if __name__ == "__main__":
    main()
